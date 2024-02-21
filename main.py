from selenium import webdriver
import time 
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC

# for xlsx
from openpyxl import load_workbook
import datetime

# current_date functions here
current_date = datetime.datetime.now()
today = current_date.strftime("%A")
print("Today is", today)

workbook =load_workbook(filename='test1.xlsx')

# days

Satur = workbook['Saturday']
Sun = workbook['Sunday']
Mon = workbook['Monday']
Tues = workbook['Tuesday']
Wednes = workbook['Wednesday']
Thurs = workbook['Thursday']
Fri = workbook['Friday']

if today == 'Saturday':
    day_name = Satur
elif today == 'Sunday':
    day_name = Sun
elif today == 'Monday':
    day_name = Mon
elif today == 'Tuesday':
    day_name = Tues
elif today == 'Wednesday':
    day_name = Wednes
elif today == 'Thursday':
    day_name = Thurs
elif today == 'Friday':
    day_name = Fri

# SELENIUM & openpyxl
    
try:
    rowno= 2
    for row in day_name.iter_rows(min_row = 3, max_row =12, min_col =3,max_col=3):
        for i in row:
            print(i.value)
            print(rowno)
            rowno +=1
            driver = webdriver.Chrome()  
            driver.maximize_window()  
            driver.get("https://www.google.com")
            element_search_box = driver.find_element(By.NAME, "q").send_keys(i.value)
            driver.find_element(By.NAME, "q").submit()
            driver.find_element(By.NAME, "q").clear()
            element_search = driver.find_element(By.NAME, "q").send_keys(i.value)
            s = [driver.find_element(By.XPATH, "//*[@id='Alh6id']/div[1]/div/ul")]
            li=[]
            lt=[]
        for i in s:
            t = i.text
            li.append(t)
            li=str(t)
            a = li.splitlines()
            lt.append(a)


        # largest  data 
        fs = 0
        lis =[]
        for j in lt:
            for y in j:
                l = len(y)
                
                if fs < l :
                    lis.clear()
                    lis.append(y)
                    
                    fs = l
                elif fs == l:
                    lis.append(y)
                    
                    fs = l
        print(lis)

        # lowest data  found 

        fs1 = fs
        lis1 =[]
        for j in lt:
            for y in j:
                l = len(y)
                
                if fs1 > l :
                    lis1.clear()
                    lis1.append(y)
                    
                    fs1 = l
                elif fs1 == l:
                    lis1.append(y)
                    
                    fs1 = l
        print(lis1)
        time.sleep(3)
        day_name.cell(row=rowno,column=4,value=lis[0])
        day_name.cell(row=rowno,column=5,value=lis1[0])
        workbook.save(filename='test1.xlsx')
        workbook.close()
        print(f'Done -- {rowno}')
except Exception as e:
    print(f'pleace colse the xlsx file and try again :  errorr : {e}')

